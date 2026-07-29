import openpyxl
import csv
import docx
from config.conexion import obtener_conexion
from limpiadores.normalizadores import comillas_sql
from inserciones.cargador_lotes import insertar_en_lotes

def cargar_staging():
    print("=" * 60)
    print("ACCIÓN 2: CARGA DE FUENTES DE DATOS CRUDAS EN STAGING")
    print("=" * 60)

    conexion = obtener_conexion()
    cursor = conexion.cursor()

    # 1. Carga desde Excel: campusfit_fuentes.xlsx
    ruta_excel = 'documentos/campusfit_fuentes.xlsx'
    print(f"Leyendo Excel '{ruta_excel}'...")
    wb = openpyxl.load_workbook(ruta_excel)

    # usuarios_crudo
    filas_u = list(wb['usuarios_crudo'].iter_rows(values_only=True))[1:]
    valores_u = [f"({','.join(comillas_sql(x) for x in r[:6])})" for r in filas_u if any(r)]
    insertar_en_lotes(cursor, 'stg_usuarios', [], valores_u)
    print(f"  - stg_usuarios: {len(valores_u)} registros.")

    # cursos_crudo
    filas_c = list(wb['cursos_crudo'].iter_rows(values_only=True))[1:]
    valores_c = [f"({','.join(comillas_sql(x) for x in r[:6])})" for r in filas_c if any(r)]
    insertar_en_lotes(cursor, 'stg_cursos', [], valores_c)
    print(f"  - stg_cursos: {len(valores_c)} registros.")

    # campanias_crudo
    filas_cmp = list(wb['campanias_crudo'].iter_rows(values_only=True))[1:]
    valores_cmp = [f"({','.join(comillas_sql(x) for x in r[:6])})" for r in filas_cmp if any(r)]
    insertar_en_lotes(cursor, 'stg_campanias', [], valores_cmp)
    print(f"  - stg_campanias: {len(valores_cmp)} registros.")

    # inscripciones_crudo
    filas_ins = list(wb['inscripciones_crudo'].iter_rows(values_only=True))[1:]
    valores_ins = [f"({','.join(comillas_sql(x) for x in r[:7])})" for r in filas_ins if any(r)]
    insertar_en_lotes(cursor, 'stg_inscripciones', [], valores_ins)
    print(f"  - stg_inscripciones: {len(valores_ins)} registros.")

    # 2. Carga desde CSV: pagos_pasarela.csv
    ruta_csv = 'documentos/pagos_pasarela.csv'
    print(f"Leyendo CSV '{ruta_csv}'...")
    with open(ruta_csv, encoding='utf-8-sig', errors='ignore') as f:
        lector = csv.reader(f)
        next(lector)
        valores_p = [f"({','.join(comillas_sql(x) for x in r[:7])})" for r in lector if any(r)]
        insertar_en_lotes(cursor, 'stg_pagos_pasarela', [], valores_p)
    print(f"  - stg_pagos_pasarela: {len(valores_p)} registros.")

    # 3. Carga desde Word: reporte_campanias.docx
    ruta_docx = 'documentos/reporte_campanias.docx'
    print(f"Leyendo Word '{ruta_docx}'...")
    doc = docx.Document(ruta_docx)
    valores_doc = []
    for tabla in doc.tables:
        for fila in tabla.rows[1:]:
            celdas = [c.text.strip() for c in fila.cells]
            if len(celdas) >= 5:
                valores_doc.append(f"({','.join(comillas_sql(x) for x in celdas[:5])})")
    insertar_en_lotes(cursor, 'stg_reporte_campanias_doc', [], valores_doc)
    print(f"  - stg_reporte_campanias_doc: {len(valores_doc)} registros.")

    conexion.close()
    print("Acción 2 finalizada exitosamente.\n")

if __name__ == "__main__":
    cargar_staging()
